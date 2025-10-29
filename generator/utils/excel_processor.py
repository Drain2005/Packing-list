
import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import shutil
import logging
from copy import copy

logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self):
        self.container_column = None
        self.template_path = None

    def set_template(self, template_path):
        self.template_path = template_path

    def read_excel_file(self, file_path):
        df = pd.read_excel(file_path, sheet_name=0)
        df.columns = df.columns.astype(str)
        df.columns = self._clean_column_names(df.columns)
        df = self._remove_duplicate_columns(df)
        if self.container_column is None:
            self.container_column = self._find_container_column(df)
        return df, list(df.columns)

    def _clean_column_names(self, columns):
        mapping = {
            'REEL NO.': 'NO_BOBINE',
            'NO BOBINE': 'NO_BOBINE',
            'NUMERO BOBINE': 'NO_BOBINE',
            'REEL_NO': 'NO_BOBINE',
            'N° BOBINE': 'NO_BOBINE',
            'CONTAINER': 'CONTENEUR',
            'CONTENEUR': 'CONTENEUR',
            'CTN': 'CONTENEUR',
            'DIAM MM': 'DIAMETRE',
            'DIAMÈTRE': 'DIAMETRE',
            'POIDS (KG)': 'POIDS',
            'PRODUCT': 'PRODUIT',
            'REF PAPIER': 'REF_PAPIER',
            'REFERENCE PAPIER': 'REF_PAPIER',
            'RÉFÉRENCE': 'REF_PAPIER',
            'METRAGE': 'METRAGE',
            'LONGUEUR': 'METRAGE'
        }
        cleaned = []
        seen = set()
        for c in columns:
            cc = mapping.get(str(c).strip().upper(), str(c).strip().upper())
            if cc in seen:
                i = 1
                while f"{cc}_{i}" in seen:
                    i += 1
                cc = f"{cc}_{i}"
            seen.add(cc)
            cleaned.append(cc)
        return cleaned

    def _remove_duplicate_columns(self, df):
        drop_cols = []
        for i, c1 in enumerate(df.columns):
            for j, c2 in enumerate(df.columns):
                if i < j and df[c1].equals(df[c2]):
                    drop_cols.append(c2)
        if drop_cols:
            df = df.drop(columns=set(drop_cols))
        return df

    def _find_container_column(self, df):
        for col in df.columns:
            if any(k in col.lower() for k in ['container', 'conteneur', 'ctn', 'cnt']):
                return col
        return df.columns[0]

    def extract_containers(self, df):
        if self.container_column in df.columns:
            return [str(c).strip() for c in df[self.container_column].dropna().unique()]
        return []
    
    def filter_by_container(self, df, container):
        if self.container_column not in df.columns:
            return df
        return df[df[self.container_column] == container]

    def _calculate_font_size(self, bobine_number):
        """Calcule la taille de police adaptative selon la longueur du numéro"""
        length = len(str(bobine_number))
        
        # Taille de police inversement proportionnelle à la longueur
        if length <= 8:
            return 20    # Grande taille pour les numéros courts
        elif length <= 12:
            return 18    # Taille moyenne
        elif length <= 16:
            return 16    # Taille réduite
        elif length <= 20:
            return 14    # Petite taille
        else:
            return 12    # Très petite taille pour les numéros très longs

    def _calculate_column_width(self, bobine_number):
        """Calcule la largeur de colonne adaptative"""
        length = len(str(bobine_number))
        
        # Largeur proportionnelle à la longueur du texte
        if length <= 8:
            return 20    # Largeur standard
        elif length <= 12:
            return 24    # Largeur moyenne
        elif length <= 16:
            return 28    # Largeur augmentée
        elif length <= 20:
            return 32    # Largeur importante
        else:
            return 38    # Très large pour les numéros très longs

    def _insert_barcode_to_excel(self, sheet, row, col, bobine_number, bobine_col):
        """Insère une FORMULE Excel pour code-barres dynamique"""
        try:
            # Créer une formule qui référence la colonne des numéros de bobine
            bobine_cell_ref = f"{openpyxl.utils.get_column_letter(bobine_col)}{row}"
            
            # La formule = simplement référencer la cellule du numéro de bobine
            formula = f"={bobine_cell_ref}"
            
            # Mettre la FORMULE dans la cellule
            cell = sheet.cell(row=row, column=col)
            cell.value = formula  # FORMULE
            
            # Calculer la taille de police adaptative basée sur le numéro actuel
            font_size = self._calculate_font_size(bobine_number)
            
            # Appliquer la police IDAutomationHC39M avec taille adaptative
            cell.font = Font(
                name='IDAutomationHC39M Free Version',
                size=font_size,
                bold=False
            )
            
            # Centrer le code-barres dans la cellule
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center'
            )
            
            # Ajuster la largeur de colonne si nécessaire
            col_letter = openpyxl.utils.get_column_letter(col)
            calculated_width = self._calculate_column_width(bobine_number)
            
            # Ne modifier la largeur que si elle est insuffisante
            current_width = sheet.column_dimensions[col_letter].width
            if current_width is None or current_width < calculated_width:
                sheet.column_dimensions[col_letter].width = calculated_width
            
            # HAUTEUR AUGMENTÉE : 45 pixels comme le template zzzz
            sheet.row_dimensions[row].height = 45.0
            
            logger.debug(f"Code-barres DYNAMIQUE inséré: formule {formula} (taille: {font_size}, hauteur: 45)")
            
        except Exception as e:
            logger.error(f"Erreur insertion code-barres dynamique: {e}")
            # Fallback: valeur statique
            sheet.cell(row=row, column=col).value = bobine_number

    def _copy_row_formatting(self, source_sheet, target_sheet, source_row, target_row):
        """Copie le formatage d'une ligne source vers une ligne cible"""
        # Copie la hauteur de ligne
        if source_row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
        
        for col in range(1, source_sheet.max_column + 1):
            source_cell = source_sheet.cell(row=source_row, column=col)
            target_cell = target_sheet.cell(row=target_row, column=col)
            
            # Copie le style
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
            
            # Copie la largeur de colonne
            col_letter = openpyxl.utils.get_column_letter(col)
            if col_letter in source_sheet.column_dimensions:
                target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

    def _add_extra_rows(self, sheet, start_row, num_extra_rows, template_row):
        """Ajoute des lignes supplémentaires en copiant le format du template"""
        # HAUTEUR AUGMENTÉE : 45 pixels comme le template zzzz
        template_height = 45.0
        
        # Insère les nouvelles lignes
        for i in range(num_extra_rows):
            sheet.insert_rows(start_row)
            
            # Appliquer la hauteur augmentée
            sheet.row_dimensions[start_row].height = template_height
            
            # Copie le format de la ligne template vers la nouvelle ligne
            self._copy_row_formatting(sheet, sheet, template_row, start_row)
        
        return start_row

    def _ensure_consistent_row_heights(self, sheet, start_row, end_row):
        """Assure que toutes les lignes ont la même hauteur augmentée"""
        # HAUTEUR AUGMENTÉE : 45 pixels pour toutes les lignes
        barcode_row_height = 45.0
        for row in range(start_row, end_row + 1):
            sheet.row_dimensions[row].height = barcode_row_height

    def create_excel(self, data, container, output_dir,
                     cariste, fournisseur, numero_dossier,
                     type_certification, numero_certificat):
        
        try:
            os.makedirs(output_dir, exist_ok=True)
            filename = f"{container}.xlsx"
            file_path = os.path.join(output_dir, filename)

            # Copie du template de base
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template introuvable : {self.template_path}")
            shutil.copy2(self.template_path, file_path)

            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook["FO57"] if "FO57" in workbook.sheetnames else workbook.active
            base_positions = self._find_field_positions(sheet)

            # Détection automatique du nombre de lignes de données dans le modèle
            start_row = base_positions.get("start_row", 15)
            
            # Trouver la dernière ligne de données dans le template
            last_data_row = start_row
            while sheet.cell(row=last_data_row, column=1).value not in (None, "", " "):
                last_data_row += 1
            
            template_data_rows = last_data_row - start_row
            logger.info(f"Template a {template_data_rows} lignes de données (de {start_row} à {last_data_row-1})")

            total_bobines = len(data)
            
            # Ajouter des lignes supplémentaires si nécessaire
            if total_bobines > template_data_rows:
                extra_rows_needed = total_bobines - template_data_rows
                logger.info(f"Ajout de {extra_rows_needed} lignes supplémentaires dans la même feuille")
                
                # Utilise la dernière ligne de données comme template pour les nouvelles lignes
                template_format_row = start_row + template_data_rows - 1
                self._add_extra_rows(sheet, start_row + template_data_rows, extra_rows_needed, template_format_row)

            #  APPLIQUER LA HAUTEUR AUGMENTÉE à toutes les lignes
            end_row = start_row + total_bobines - 1
            self._ensure_consistent_row_heights(sheet, start_row, end_row)

            # Remplir toutes les données dans la même feuille
            self._fill_single_sheet(sheet, data, base_positions,
                                   container, cariste, fournisseur, numero_dossier,
                                   type_certification, numero_certificat)

            # Renommer la feuille principale
            sheet.title = "FO57"

            # Supprimer les autres feuilles inutiles si elles existent
            sheets_to_remove = [name for name in workbook.sheetnames if name != "FO57" and name != "Mode de remplisage"]
            for sheet_name in sheets_to_remove:
                del workbook[sheet_name]

            workbook.save(file_path)
            logger.info(f"Fichier {file_path} créé avec {total_bobines} bobines, hauteur augmentée (45px)")
            return file_path

        except Exception as e:
            logger.error(f"Erreur lors de la création Excel : {e}", exc_info=True)
            raise

    def _fill_single_sheet(self, sheet, data, positions, container, cariste, fournisseur,
                          numero_dossier, type_certification, numero_certificat):
        """Remplit une seule feuille avec toutes les données"""
        # Mettre à jour les en-têtes
        if 'container' in positions:
            sheet[positions['container']] = f"N° CT : {container}"
        if 'cariste' in positions:
            sheet[positions['cariste']] = f"CARISTE : {cariste}"
        if 'date' in positions:
            sheet[positions['date']] = f"DATE : {datetime.now().strftime('%d/%m/%Y')}"
        if 'dossier' in positions:
            sheet[positions['dossier']] = f"No. Dossier : {numero_dossier}"

        start_row = positions.get("start_row", 15)
        code_barre_col = positions.get('col_code_barre', 4)  # Colonne D par défaut
        bobine_col = positions.get('col_bobine', 2)  # Colonne B par défaut

        for idx, (_, row) in enumerate(data.iterrows(), 1):
            excel_row = start_row + idx - 1
            bobine_number = row.get('NO_BOBINE', '')

            if 'col_numero' in positions:
                sheet.cell(row=excel_row, column=positions['col_numero'], value=idx)
            if 'col_bobine' in positions:
                sheet.cell(row=excel_row, column=positions['col_bobine'], value=bobine_number)
            if 'col_fournisseur' in positions:
                sheet.cell(row=excel_row, column=positions['col_fournisseur'], value=fournisseur)
            if 'col_reference' in positions:
                sheet.cell(row=excel_row, column=positions['col_reference'], value=row.get('REF_PAPIER', ''))
            if 'col_diametre' in positions:
                sheet.cell(row=excel_row, column=positions['col_diametre'], value=row.get('DIAMETRE', ''))
            if 'col_poids' in positions:
                sheet.cell(row=excel_row, column=positions['col_poids'], value=row.get('POIDS', ''))
            if 'col_certificat' in positions:
                sheet.cell(row=excel_row, column=positions['col_certificat'], value=str(numero_certificat))
            if 'col_type_certif' in positions:
                sheet.cell(row=excel_row, column=positions['col_type_certif'], value=type_certification)

            # Vider le contenu existant de la cellule code-barre
            sheet.cell(row=excel_row, column=code_barre_col).value = None
            
            # Insérer une FORMULE dynamique au lieu d'une valeur fixe
            if bobine_number and bobine_number not in ['', 'NaN', 'None']:
                self._insert_barcode_to_excel(sheet, excel_row, code_barre_col, bobine_number, bobine_col)
            else:
                # Si pas de numéro de bobine, laisser une formule vide
                sheet.cell(row=excel_row, column=code_barre_col).value = ""

    def _find_field_positions(self, sheet):
        """Détecte la position des champs dans le template FO57"""
        pos = {}
        for row in sheet.iter_rows(max_row=20):
            for cell in row:
                if cell.value:
                    val = str(cell.value).upper().strip()
                    if "CARISTE" in val:
                        pos["cariste"] = cell.coordinate
                    if "N° CT" in val or "CONTENEUR" in val:
                        pos["container"] = cell.coordinate
                    if "DATE" in val:
                        pos["date"] = cell.coordinate
                    if "DOSSIER" in val:
                        pos["dossier"] = cell.coordinate

        for r in range(1, 40):
            for c in range(1, 40):
                v = str(sheet.cell(row=r, column=c).value or "").upper().strip()
                if v in ["N°", "NO", "NUMERO"]:
                    pos["col_numero"] = c
                    pos["start_row"] = r + 1
                elif "N° FOURNISSEUR" in v or "NO FOURNISSEUR" in v:
                    pos["col_bobine"] = c
                elif v == "FOURNISSEUR":
                    pos["col_fournisseur"] = c
                elif "REF" in v or "RÉFÉRENCE" in v:
                    pos["col_reference"] = c
                elif "DIAM" in v:
                    pos["col_diametre"] = c
                elif "POIDS" in v:
                    pos["col_poids"] = c
                elif v in ["N° CERTIFICAT FSC", "CERTIFICAT FSC"]:
                    pos["col_certificat"] = c
                elif "TYPE" in v or "CERTIFICATION" in v:
                    pos["col_type_certif"] = c
                elif "CODE BARRE" in v:
                    pos["col_code_barre"] = c
        return pos